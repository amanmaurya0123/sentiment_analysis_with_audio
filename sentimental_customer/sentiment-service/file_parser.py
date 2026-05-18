import fitz
from PIL import Image, ImageOps
import pytesseract
from docx import Document
from openpyxl import load_workbook
from faster_whisper import WhisperModel
import os

whisper_model = None


pytesseract.pytesseract.tesseract_cmd = (
    r"C:\Users\Aman.Maurya\AppData\Local\Programs\Tesseract-OCR\tesseract.exe"
)

SUPPORTED_EXTENSIONS = {
    ".pdf",
    ".doc",
    ".docx",
    ".txt",
    ".png",
    ".jpg",
    ".jpeg",
    ".webp",
    ".csv"
}

def parse_file(filepath: str) -> str:
    """
    Reads a file and returns its full text content as a single string.
    Supports PDF, DOCX, TXT, CSV, XLSX.
    """
    _, ext = os.path.splitext(filepath.lower())

    if ext not in SUPPORTED_EXTENSIONS:
        raise ValueError(f"Unsupported file type: {ext}. Supported: {', '.join(SUPPORTED_EXTENSIONS)}")

    if ext == ".pdf":
        return _parse_pdf(filepath)
    elif ext == ".docx":
        return _parse_docx(filepath)
    elif ext in (".txt", ".csv"):
        return _parse_text(filepath)
    elif ext in (".png", ".jpg", ".jpeg", ".webp"):
        return _parse_image(filepath)


def _parse_pdf(filepath: str) -> str:
    doc   = fitz.open(filepath)
    pages = [page.get_text() for page in doc]
    doc.close()
    return "\n".join(pages).strip()


def _parse_docx(filepath: str) -> str:
    doc   = Document(filepath)
    lines = [para.text for para in doc.paragraphs if para.text.strip()]
    return "\n".join(lines).strip()


# def _parse_xlsx(filepath: str) -> str:
#     wb    = load_workbook(filepath, read_only=True, data_only=True)
#     lines = []
#     for sheet in wb.worksheets:
#         for row in sheet.iter_rows(values_only=True):
#             line = " | ".join(str(cell) for cell in row if cell is not None)
#             if line.strip():
#                 lines.append(line)
#     wb.close()
#     return "\n".join(lines).strip()


def _parse_text(filepath: str) -> str:
    with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
        return f.read().strip()


def _parse_image(filepath: str) -> str:

    with Image.open(filepath) as image:
        image = ImageOps.exif_transpose(image)
        if image.mode not in ("RGB", "L"):
            image = image.convert("RGB")
        text = pytesseract.image_to_string(image)

    return text.strip()


def split_into_chunks(text: str, max_chars: int = 500) -> list[str]:
    """
    Splits text into sentence-aware chunks for per-chunk analysis.
    Tries to split on sentence boundaries first, then falls back to char limit.
    """
    import re
    # Split on sentence endings
    sentences = re.split(r'(?<=[.!?])\s+', text)
    chunks  = []
    current = ""

    for sentence in sentences:
        if not sentence.strip():
            continue
        if len(current) + len(sentence) <= max_chars:
            current += (" " if current else "") + sentence
        else:
            if current:
                chunks.append(current.strip())
            if len(sentence) > max_chars:
                for i in range(0, len(sentence), max_chars):
                    chunks.append(sentence[i:i + max_chars].strip())
            else:
                current = sentence

    if current.strip():
        chunks.append(current.strip())

    return [c for c in chunks if c]

model = WhisperModel("base", device="cpu", compute_type="int8")

def speech_to_text(audio_path):
    segments, info = model.transcribe(audio_path)
    text = ""
    for segment in segments:
        text += segment.text
    return text.strip()